from openpyxl import Workbook, load_workbook
import pandas as pd
import csv
import os

individualsnf = [0]
individualsf = [0]
individualspet = [0]
individualsgr = [0]
equipsnf = [0]
equipsf = [0]

# indnf = []
# indnfnames = []
# indnfteams = []
# indnfmitjana = []
# indnfbitlles = []
# indnfpunts = []
# indnftirades = []
# indnfdilluns = []
# indnfdimarts = []
# indnfdimecres = []
# indnfdijous = []
# indnfdivendres = []

# indfnames = []
# indfteams = []
# indfmitjana = []
# indfbitlles = []
# indfpunts = []
# indftirades = []
# indfdilluns = []
# indfdimarts = []
# indfdimecres = []
# indfdijous = []
# indfdivendres = []

# indpetnames = []
# indpetteams = []
# indpetmitjana = []
# indpetbitlles = []
# indpetpunts = []
# indpettirades = []
# indpetdilluns = []
# indpetdimarts = []
# indpetdimecres = []
# indpetdijous = []
# indpetdivendres = []

# indgrnames = []
# indgrteams = []
# indgrmitjana = []
# indgrbitlles = []
# indgrpunts = []
# indgrtirades = []
# indgrdilluns = []
# indgrdimarts = []
# indgrdimecres = []
# indgrdijous = []
# indgrdivendres = []

# eqnfteams = []
# eqnfdilluns = []
# eqnfdimarts = []
# eqnfdimecres = []
# eqnfdijous = []
# eqnfdivendres = []
# eqnftotpunts = []
# eqnfbitlles = []

# eqfteams = []
# eqfdilluns = []
# eqfdimarts = []
# eqfdimecres = []
# eqfdijous = []
# eqfdivendres = []
# eqftotpunts = []
# eqfbitlles = []


def csvexec(inf=individualsnf, indf=individualsf, ipet=individualspet, igr=individualsgr, ef=equipsf, enf=equipsnf):
	global tot
	tot = [0]
	
	if os.path.exists('Web.csv'):
		os.remove('Web.csv')
		
	read_file = pd.read_excel("Web.xlsx")
	read_file.to_csv("Web.csv",
                  index = None,
                  header=None)

	filename = 'Web.xlsx'
	
	workbook = load_workbook(filename, data_only=True, read_only=True)
	sheet = workbook.active
	
	inf.insert(0, sheet["A1"].value)
	indf.insert(0, sheet["B1"].value)
	ipet.insert(0, sheet["C1"].value)
	igr.insert(0, sheet["D1"].value)
	enf.insert(0, sheet["E1"].value)
	ef.insert(0, sheet["F1"].value)

	with open('Web.csv', mode ='r') as file_obj:
		csvfile = csv.reader(file_obj)
		for lines in csvfile:
			tot.append(lines)
		print(tot)


# def exec(inf=individualsnf, indf=individualsf, ipet=individualspet, igr=individualsgr, ef=equipsf, enf=equipsnf):

# 	#WE LOAD THE WB EVERY TIME SO IT GETS UPDATED WITH THE NEW RESULTS
# 	filename = 'Web.xlsx'
	
# 	workbook = load_workbook(filename, data_only=True, read_only=True)
# 	sheet = workbook.active

# 	#GET THE NUMBER OF CANDIDATES/TEAMS
# 	inf.insert(0, sheet["A1"].value)
# 	indf.insert(0, sheet["B1"].value)
# 	ipet.insert(0, sheet["C1"].value)
# 	igr .insert(0, sheet["D1"].value)
# 	enf.insert(0, sheet["E1"].value)
# 	ef.insert(0, sheet["F1"].value)
# 	#GET THE NUMBER OF CANDIDATES/TEAMS


# 	#GET THE NAMES OF THE INDNF CATEGORY
	
# 	infi = 1 #Starting pos. is on one to match sheet
	
# 	for infi in range(inf[0]):

# 		c1 = "B" + str(infi+1)
# 		c2 = "C" + str(infi+1)
# 		c3 = "D" + str(infi+1)
# 		c4 = "E" + str(infi+1)
# 		c5 = "F" + str(infi+1)
# 		c6 = "G" + str(infi+1)
# 		c7 = "H" + str(infi+1)
# 		c8 = "I" + str(infi+1)
# 		c9 = "J" + str(infi+1)
# 		c10 = "K" + str(infi+1)
# 		c11 = "L" + str(infi+1)
		
# 		indnfnames.append(sheet[c1].value)
# 		indnfteams.append(sheet[c2].value)
# 		indnfmitjana.append(sheet[c3].value)
# 		indnfbitlles.append(sheet[c4].value)
# 		indnfpunts.append(sheet[c5].value)
# 		indnftirades.append(sheet[c6].value)
# 		indnfdilluns.append(sheet[c7].value)
# 		indnfdimarts.append(sheet[c8].value)
# 		indnfdimecres.append(sheet[c9].value)
# 		indnfdijous.append(sheet[c10].value)
# 		indnfdivendres.append(sheet[c11].value)

# 		print(infi)
# 	print('Infi done')

# 	#GET THE NAMES OF THE INDNF CATEGORY

# 	#GET THE NAMES OF THE F CATEGORY
	
# 	indfi = inf[0] + 1 
	
# 	for indfi in range(inf[0]+indf[0]):
# 		c1 = "B" + str(indfi+1)
# 		c2 = "C" + str(indfi+1)
# 		c3 = "D" + str(indfi+1)
# 		c4 = "E" + str(indfi+1)
# 		c5 = "F" + str(indfi+1)
# 		c6 = "G" + str(indfi+1)
# 		c7 = "H" + str(indfi+1)
# 		c8 = "I" + str(indfi+1)
# 		c9 = "J" + str(indfi+1)
# 		c10 = "K" + str(indfi+1)
# 		c11 = "L" + str(indfi+1)
	
# 		indfnames.append(sheet[c1].value)
# 		indfteams.append(sheet[c2].value)
# 		indfmitjana.append(sheet[c3].value)
# 		indfbitlles.append(sheet[c4].value)
# 		indfpunts.append(sheet[c5].value)
# 		indftirades.append(sheet[c6].value)
# 		indfdilluns.append(sheet[c7].value)
# 		indfdimarts.append(sheet[c8].value)
# 		indfdimecres.append(sheet[c9].value)
# 		indfdijous.append(sheet[c10].value)
# 		indfdivendres.append(sheet[c11].value)
	
# 	#GET THE NAMES OF THE F CATEGORY

# 	#GET THE NAMES OF THE PET CATEGORY
	
# 	indpeti = inf[0]+indf[0] + 1 
	
# 	for indpeti in range(inf[0]+indf[0]+ipet[0]):
# 		c1 = "B" + str(indpeti+1)
# 		c2 = "C" + str(indpeti+1)
# 		c3 = "D" + str(indpeti+1)
# 		c4 = "E" + str(indpeti+1)
# 		c5 = "F" + str(indpeti+1)
# 		c6 = "G" + str(indpeti+1)
# 		c7 = "H" + str(indpeti+1)
# 		c8 = "I" + str(indpeti+1)
# 		c9 = "J" + str(indpeti+1)
# 		c10 = "K" + str(indpeti+1)
# 		c11 = "L" + str(indpeti+1)
	
# 		indpetnames.append(sheet[c1].value)
# 		indpetteams.append(sheet[c2].value)
# 		indpetmitjana.append(sheet[c3].value)
# 		indpetbitlles.append(sheet[c4].value)
# 		indpetpunts.append(sheet[c5].value)
# 		indpettirades.append(sheet[c6].value)
# 		indpetdilluns.append(sheet[c7].value)
# 		indpetdimarts.append(sheet[c8].value)
# 		indpetdimecres.append(sheet[c9].value)
# 		indpetdijous.append(sheet[c10].value)
# 		indpetdivendres.append(sheet[c11].value)
	
# 	#GET THE NAMES OF THE PET CATEGORY

# 	#GET THE NAMES OF THE GR CATEGORY
	
# 	indgri = inf[0]+indf[0]+ipet[0] + 1 
	
# 	for indgri in range(inf[0]+indf[0]+ipet[0]+igr[0]):
# 		c1 = "B" + str(indgri+1)
# 		c2 = "C" + str(indgri+1)
# 		c3 = "D" + str(indgri+1)
# 		c4 = "E" + str(indgri+1)
# 		c5 = "F" + str(indgri+1)
# 		c5 = "G" + str(indgri+1)
# 		c6 = "G" + str(indgri+1)
# 		c7 = "H" + str(indgri+1)
# 		c8 = "I" + str(indgri+1)
# 		c9 = "J" + str(indgri+1)
# 		c10 = "K" + str(indgri+1)
# 		c11 = "L" + str(indgri+1)
		
	
# 		indgrnames.append(sheet[c1].value)
# 		indgrteams.append(sheet[c2].value)
# 		indgrmitjana.append(sheet[c3].value)
# 		indgrbitlles.append(sheet[c4].value)
# 		indgrpunts.append(sheet[c5].value)
# 		indgrtirades.append(sheet[c6].value)
# 		indfdilluns.append(sheet[c7].value)
# 		indgrdimarts.append(sheet[c8].value)
# 		indgrdimecres.append(sheet[c9].value)
# 		indgrdijous.append(sheet[c10].value)
# 		indgrdivendres.append(sheet[c11].value)
		
# 	#GET THE NAMES OF THE GR CATEGORY

# 	#GET THE TEAMS INFO ON THE NF CATEGORY

# 	eqnfi = inf[0]+indf[0]+ipet[0]+igr[0] + 1 


# 	for eqnfi in range(inf[0]+indf[0]+ipet[0]+igr[0]+enf[0]):
# 		c1 = "B" + str(eqnfi+1)
# 		c2 = "C" + str(eqnfi+1)
# 		c3 = "D" + str(eqnfi+1)
# 		c4 = "E" + str(eqnfi+1)
# 		c5 = "F" + str(eqnfi+1)
# 		c6 = "G" + str(eqnfi+1)
# 		c7 = "H" + str(eqnfi+1)
# 		c8 = "I" + str(eqnfi+1)
	
# 		eqnfteams.append(sheet[c1].value)
# 		eqnfdilluns.append(sheet[c2].value)
# 		eqnfdimarts.append(sheet[c3].value)
# 		eqnfdimecres.append(sheet[c4].value)
# 		eqnfdijous.append(sheet[c5].value)
# 		eqnfdivendres.append(sheet[c6].value)
# 		eqnftotpunts.append(sheet[c7].value)
# 		eqnfbitlles.append(sheet[c8].value)
	
# 	#GET THE TEAMS INFO ON THE NF CATEGORY
	
# 	#GET THE TEAMS INFO ON THE F CATEGORY

# 	eqfi = inf[0]+indf[0]+ipet[0]+igr[0]+enf[0] + 1 



# 	for eqfi in range(inf[0]+indf[0]+ipet[0]+igr[0]+enf[0]+ef[0]):
# 		c1 = "B" + str(eqfi+1)
# 		c2 = "C" + str(eqfi+1)
# 		c3 = "D" + str(eqfi+1)
# 		c4 = "E" + str(eqfi+1)
# 		c5 = "F" + str(eqfi+1)
# 		c6 = "G" + str(eqfi+1)
# 		c7 = "H" + str(eqfi+1)
# 		c8 = "I" + str(eqfi+1)
	
# 		eqfteams.append(sheet[c1].value)
# 		eqfdilluns.append(sheet[c2].value)
# 		eqfdimarts.append(sheet[c3].value)
# 		eqfdimecres.append(sheet[c4].value)
# 		eqfdijous.append(sheet[c5].value)
# 		eqfdivendres.append(sheet[c6].value)
# 		eqftotpunts.append(sheet[c7].value)
# 		eqfbitlles.append(sheet[c8].value)
		
# 	#GET THE TEAMS INFO ON THE F CATEGORY
# 	print('sheetoperator Done')
# 	return("Done")


def check():
	print (equipsnf)